import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
import os
import requests
from io import BytesIO


# # Verificar el directorio de trabajo actual
# st.write("Directorio de trabajo actual:", os.getcwd())

# # Cargar la plantilla de Excel
# def load_workbook():
#     return openpyxl.load_workbook('C:/Users/Usuario/Documents/Proyectos Subtec/Herramienta Planeación/Herramienta Planeación/template.xlsx')

# # Guardar datos en celdas específicas
# def save_to_excel(name, age, gender):
#     wb = load_workbook()
#     sheet = wb.active
#     sheet['A1'] = name  # Guardar nombre en celda A1
#     sheet['B1'] = age   # Guardar edad en celda B1
#     sheet['C1'] = gender # Guardar género en celda C1
#     wb.save('output.xlsx')

# # Crear la interfaz de usuario
# st.title("Formulario de Información")

# name = st.text_input("Nombre")
# age = st.number_input("Edad", min_value=0, max_value=120, step=1)
# gender = st.selectbox("Género", ["Masculino", "Femenino", "Otro"])

# if st.button("Guardar"):
#     save_to_excel(name, age, gender)
#     st.success("Datos guardados en output.xlsx")



# URL del archivo en GitHub (reemplaza con tu propia URL)
url = 'https://github.com/Manu-HR88/Prueba-1/blob/main/template.xlsx'

# Descargar el archivo desde GitHub
@st.cache
def download_template(url):
    response = requests.get(url)
    response.raise_for_status()  # Asegurarse de que la solicitud fue exitosa
    return BytesIO(response.content)

# Cargar la plantilla de Excel desde el archivo descargado
def load_workbook():
    file_stream = download_template(url)
    return openpyxl.load_workbook(file_stream)

# Guardar datos en celdas específicas
def save_to_excel(name, age, gender):
    wb = load_workbook()
    sheet = wb.active
    sheet['A1'] = name  # Guardar nombre en celda A1
    sheet['B1'] = age   # Guardar edad en celda B1
    sheet['C1'] = gender # Guardar género en celda C1
    wb.save('output.xlsx')

# Crear la interfaz de usuario
st.title("Formulario de Información")

name = st.text_input("Nombre")
age = st.number_input("Edad", min_value=0, max_value=120, step=1)
gender = st.selectbox("Género", ["Masculino", "Femenino", "Otro"])

if st.button("Guardar"):
    save_to_excel(name, age, gender)
    st.success("Datos guardados en output.xlsx")
